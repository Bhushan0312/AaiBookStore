
from tkinter import *
from tkinter import messagebox
import openpyxl
from random import randint
class new_order():
    def __init__(self,root,xl):
        self.xl= xl
        self.xl2=openpyxl.load_workbook("Resourse/Order.xlsx",data_only=True)
        self.root1 = root
        self.mainfunc()
        self.a=0
        self.b=0
    def clear(self):
        try:
            self.Cancel.destroy()
            if self.a != 0:
                for i in range(0, self.a):
                    try:
                        self.sub[i].destroy()
                        self.Qntt[i].destroy()
                        self.Qtext[i].destroy()
                    except:
                        pass
                    try:
                        self.med_b[i].destroy()
                        self.comN[i].destroy()
                        self.bookN[i].destroy()
                    except:
                        pass


            try:
                self.del_or.destroy()
            except:
                pass
            try:
                self.submit.destroy()
            except:
                pass
            try:
                self.sl.destroy()
                self.qn.destroy()
                self.oqn.destroy()
                self.Final.destroy()
                self.save_bt.destroy()
                self.Cancel.destroy()
            except:
                pass
            try:
                self.drop_medium.destroy()
                self.drop_bookcn.destroy()
                self.drop_bookbn.destroy()
                print("book removed")
                self.drop_std.destroy()
                self.stdL.destroy()
                self.stdM.destroy()
                self.stdc.destroy()
                self.stdB.destroy()
            except:
                pass
            try:
                self.sl.destroy()
                self.md.destroy()
                self.con.destroy()
                self.BN.destroy()
                self.oqn.destroy()
            except:
                pass
        except:
            pass
    def delete(self):
        self.clear()
        if(len(self.search_Text.get())!=0):
            if str(self.search_Text.get()) in self.xl2.sheetnames:
                sh2=self.xl2["Main_Entry"]
                for i in range(2,sh2.max_row+1):
                    if str(self.search_Text.get())==sh2.cell(i,1).value:
                        sh2.delete_rows(i)
                        order = self.xl2.get_sheet_by_name(str(self.search_Text.get()))
                        self.xl2.remove_sheet(order)
                        self.xl2.save("Resourse/Order.xlsx")
                        self.xl2.close()
                        self.xl2 = openpyxl.load_workbook("Resourse/Order.xlsx")
                        self.del_or = Label(self.root1, text=str(self.search_Text.get()) + " Order Deleted",font=("Arial", 14, "bold"), bg='#FFFFFF', fg='green')
                        self.del_or.place(x=450, y=300)
                    else:
                        pass
            else:
                pass
    def open(self):
        self.cancel()
        self.clear()
        if (len(self.search_Text.get()) != 0):
            if str(self.search_Text.get()) in self.xl2.sheetnames:
                sh2 = self.xl2["Main_Entry"]
                for i in range(2,sh2.max_row+1):
                    if str(sh2.cell(i,1).value)==self.search_Text.get():
                        sh2=self.xl2[sh2.cell(i,1).value]
                        standard_options=[""]
                        for i in range(2,sh2.max_row+1):
                            standard_options.append(sh2.cell(i,1).value)
                        standard_options=list(dict.fromkeys(standard_options))
                        print(standard_options)
                        click_std=StringVar()
                        click_std.set("")
                        self.stdL = Label(self.root1, text="Standard", font=("Arial", 10, "bold"), bg='#FFFFFF')
                        self.stdL.place(x=410, y=180)
                        self.drop_std = OptionMenu(self.root1, click_std, *standard_options, command=self.order_data)
                        self.drop_std.config(width=3)
                        self.drop_std.place(x=410,y=200)
                        self.edit= Button(self.root1,text='Edit',command=self.edit)
                        self.edit.place(x=700,y=200)

            else:
                self.submit = Label(self.root1, text="Order Not Fond", font=("Arial", 14, "bold"), bg='#FFFFFF',fg='red')
                self.submit.place(x=450, y=300)

    def edit(self):
        self.b=1
        sh2 = self.xl2[self.search_Text.get()]
        self.Qtext=[]
        self.a=0
        print(sh2.max_row)
        for i in range(2, sh2.max_row + 1):
                if sh2.cell(i,1).value==self.edit1:
                    self.Qtext.append(Entry(self.root1, width=3, bd=3))
                    self.Qtext[self.a].place(x=700, y=(250 + (self.a * 20)))
                    self.a+=1
        self.save_bt = Button(self.root1, text='Update Order', command=self.Save_or)
        self.save_bt.place(x=730, y=250 + (self.a * 22))
        self.Cancel = Button(self.root1, text='Cancel', command=self.cancel)
        self.Cancel.place(x=840, y=250 + (self.a * 22))


    def order_data(self,event):
        self.edit1=event
        try:
            self.sl.destroy()
            self.md.destroy()
            self.con.destroy()
            self.BN.destroy()
            self.oqn.destroy()
        except:
            pass
        sh2=self.xl2[self.search_Text.get()]
        if self.a!=0:
            for i in range(0,self.a):
                try:
                    self.sub[i].destroy()
                    self.med_b[i].destroy()
                    self.comN[i].destroy()
                    self.bookN[i].destroy()
                    self.Qntt[i].destroy()
                except:
                    pass
        else:
            pass

        self.sub = []
        self.med_b = []
        self.comN=[]
        self.bookN=[]
        self.Qntt = []
        self.a=0
        self.sl = Label(self.root1, text="Subject", font=("Arial", 10, "bold"), bg='#FFFFFF', fg='green')
        self.sl.place(x=410, y=230)
        self.md = Label(self.root1, text="medium", font=("Arial", 10, "bold"), bg='#FFFFFF', fg='green')
        self.md.place(x=560, y=230)
        self.con = Label(self.root1, text="com.Name", font=("Arial", 10, "bold"), bg='#FFFFFF', fg='green')
        self.con.place(x=630, y=230)
        self.BN = Label(self.root1, text="BookName", font=("Arial", 10, "bold"), bg='#FFFFFF', fg='green')
        self.BN.place(x=730, y=230)
        self.oqn = Label(self.root1, text="Quantity", font=("Arial", 10, "bold"), bg='#FFFFFF', fg='green')
        self.oqn.place(x=820, y=230)
        for i in range(2, sh2.max_row + 1):
            if (event== sh2.cell(i,1).value):
                self.sub.append(Label(self.root1, text=sh2.cell(i, 3).value, font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.sub[self.a].place(x=410, y=(250 + (self.a * 20)))
                self.med_b.append(Label(self.root1, text=sh2.cell(i, 2).value, font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.med_b[self.a].place(x=570, y=(250 + (self.a * 20)))
                self.comN.append(Label(self.root1, text=sh2.cell(i, 4).value, font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.comN[self.a].place(x=630, y=(250 + (self.a * 20)))
                self.bookN.append(Label(self.root1, text=sh2.cell(i, 5).value, font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.bookN[self.a].place(x=740, y=(250 + (self.a * 20)))
                self.Qntt.append(Label(self.root1, text=sh2.cell(i, 6).value, font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.Qntt[self.a].place(x=840, y=(250 + (self.a * 20)))
                self.a+=1
            else:
                pass


    def cancel(self):
        self.clear()
        try:
            print(self.newID)
            if self.newID in self.xl2.sheetnames:
                print(self.newID)
                order=self.xl2.get_sheet_by_name(self.newID)
                self.xl2.remove_sheet(order)
                sh=self.xl2["Main_Entry"]
                sh.cell(sh.max_row,1).value=None
                self.xl2.save("Resourse/Order.xlsx")
                self.xl2.close()
                self.xl2 = openpyxl.load_workbook("Resourse/Order.xlsx")
            else:
                pass
        except:
            pass

    def Warning(self):
        MsgBox = messagebox.askquestion('Warning', 'Are you sure ? ',icon='warning')
        if MsgBox == 'yes':
            self.delete()
        else:
            pass


    def random_ID(self):
        range_start = 10 ** (8 - 1)
        range_end = (10 ** 8) - 1
        return randint(range_start, range_end)

    def Save_or(self):
        for i in range(1,self.a+1):
            if (len(self.Qtext[i-1].get()) != 0 ):
                try:
                    sh2 = self.xl2[self.newID]
                except:
                    sh2 = self.xl2[self.search_Text.get()]
                if (self.med == sh2.cell(i, 1).value and self.cn == sh2.cell(i, 2).value and self.blist == sh2.cell(i,4).value):
                    if (len(self.search_Text.get()) != 0 or len(self.Qtext[i-1].get()) != 0 ):
                        if str(self.search_Text.get()) in self.xl2.sheetnames:
                            sh2.cell(i, 6).value = self.Qtext[i].get()
                        else:
                            sh2.cell(i, 6).value = self.Qtext[i].get()
                else:
                    row = sh2.max_row + 1
                    print(self.sub_dt[i-1])
                    sh2.cell(row, 1).value = self.click_std.get()
                    sh2.cell(row, 2).value = self.med
                    sh2.cell(row, 3).value = self.sub_dt[i-1]
                    sh2.cell(row, 4).value = self.cn
                    sh2.cell(row, 5).value = self.blist
                    sh2.cell(row, 6).value = self.Qtext[i-1].get()
                    self.xl2.save("Resourse/Order.xlsx")

    def set_order(self):
        self.Save_or()
        self.xl2.save("Resourse/Order.xlsx")
        self.xl2.close()
        self.xl2=openpyxl.load_workbook("Resourse/Order.xlsx")
        self.clear()
        self.submit=Label(self.root1,text=self.newID+" Order placed successfully",font=("Arial", 14, "bold"), bg='#FFFFFF',fg='green')
        self.submit.place(x=450,y=300)
        self.newID = 0



    def medium(self,event):
            self.drop_medium.destroy()
            sh = self.xl[event]
            self.medium_options = [""]
            for i in range(2, (sh.max_row) + 1):
                self.medium_options.append(sh.cell(i, 1).value)
            self.medium_options = list(dict.fromkeys(self.medium_options))
            self.medium_options.reverse()
            self.click_mdm = StringVar()
            self.click_mdm.set("")
            self.drop_medium = OptionMenu(self.root1, self.click_mdm, *self.medium_options, command=self.company_n)
            self.drop_medium.config(width=10)
            self.drop_medium.place(x=500, y=200)
    def company_n(self,event):
        self.drop_bookcn.destroy()
        self.med = event
        sh = self.xl[self.click_std.get()]
        cn = [""]
        for i in range(2, (sh.max_row) + 1):
            cn.append(sh.cell(i, 2).value)
        cn = list(dict.fromkeys(cn))
        self.click_cn = StringVar()
        self.click_cn.set("")
        self.drop_bookcn = OptionMenu(self.root1, self.click_cn, *cn, command=self.bookname)
        self.drop_bookcn.config(width=8)
        self.drop_bookcn.place(x=620, y=200)
    def bookname(self,event):
        self.drop_bookbn.destroy()
        self.cn = event
        sh = self.xl[self.click_std.get()]
        blist = [""]
        for i in range(2, (sh.max_row) + 1):
            if event == sh.cell(i, 2).value:
                blist.append(sh.cell(i, 4).value)
            else:
                pass
        blist = list(dict.fromkeys(blist))
        click_cn = StringVar()
        click_cn.set("")
        self.drop_bookbn = OptionMenu(self.root1, click_cn,*blist, command=self.book_select)
        self.drop_bookbn.config(width=8)
        self.drop_bookbn.place(x=750, y=200)
        print("book_created")
    def book_select(self,event):
        self.blist=event
        sh=self.xl[self.click_std.get()]
        if self.a!=0:
            for i in range(0,self.a):
                self.sub[i].destroy()
                self.Qntt[i].destroy()
                self.Qtext[i].destroy()
        try:
            self.Cancel.destroy()
            self.sl.destroy()
            self.qn.destroy()
            self.oqn.destroy()
            self.Final.destroy()
            self.save_bt.destroy()
        except:
            pass
        self.sub = []
        self.sub_dt=[]
        self.Qntt = []
        self.Qtext=[]
        self.a=0
        self.sl=Label(self.root1, text="Subject" , font=("Arial", 10, "bold"), bg='#FFFFFF',fg='green')
        self.sl.place(x=410,y=230)
        self.qn = Label(self.root1, text="Current Quantity", font=("Arial", 10, "bold"), bg='#FFFFFF',fg='green')
        self.qn.place(x=540, y=230)
        self.oqn = Label(self.root1, text="Order Quantity", font=("Arial", 10, "bold"), bg='#FFFFFF', fg='green')
        self.oqn.place(x=670, y=230)
        for i in range(2,sh.max_row+1):
            if (self.med==sh.cell(i,1).value and self.cn==sh.cell(i,2).value and self.blist==sh.cell(i,4).value):
                self.sub_dt.append(sh.cell(i, 3).value)
                self.sub.append(Label(self.root1, text=self.sub_dt[self.a], font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.sub[self.a].place(x=410, y=(250 +(self.a*20)))
                self.Qntt.append(Label(self.root1, text=sh.cell(i, 5).value, font=("Arial", 9, "bold"), bg='#FFFFFF'))
                self.Qntt[self.a].place(x=590, y=(250 + (self.a * 20)))
                self.Qtext.append(Entry(self.root1,width=3,bd=3))
                self.Qtext[self.a].place(x=700,y=(250 +(self.a*20)))
                self.a+=1

        self.save_bt = Button(self.root1, text='Save',command=self.Save_or)
        self.save_bt.place(x=730, y=250 + (self.a * 22))
        self.Final=Button(self.root1,text='Set Order',command=self.set_order)
        self.Final.place(x=770,y=250+(self.a*22))
        self.Cancel=Button(self.root1,text='Cancel',command=self.cancel)
        self.Cancel.place(x=840, y=250 + (self.a*22))
    def All_sub(self,event):
        pass

    def new_or(self):
        self.clear()
        sh2=self.xl2["Main_Entry"]
        n1=self.xl2.create_sheet()
        self.newID= str(self.random_ID())
        #self.newID='41487996'
        sh2.cell(sh2.max_row + 1, 1).value = self.newID
        for i in range(2,sh2.max_row+1):
            if sh2.cell(i,1).value==self.newID:
                while True:
                    if self.newID in self.xl2.sheetnames:
                        print(self.newID)
                        sh2.cell(sh2.max_row, 1).value = ''
                        self.newID = str(self.random_ID())
                        sh2.cell(sh2.max_row, 1).value = self.newID
                    else:
                        break
            else:
                pass

        n1.title=self.newID
        self.xl2.save("Resourse/Order.xlsx")
        self.xl2.close()
        self.xl2=openpyxl.load_workbook("Resourse/Order.xlsx")
        self.a=0
        standard_options = [""]
        for i in self.xl.sheetnames:
            standard_options.append(i)
        self.click_std = StringVar()
        self.click_std.set(standard_options[0])
        self.stdL = Label(self.root1, text="Standard", font=("Arial", 10, "bold"), bg='#FFFFFF')
        self.stdL.place(x=410, y=180)
        self.drop_std = OptionMenu(self.root1, self.click_std, *standard_options,command=self.medium)
        self.drop_std.config(width=3)
        self.drop_std.place(x=410, y=200)
        self.stdM = Label(self.root1, text="Medium/Stream", font=("Arial", 10, "bold"), bg='#FFFFFF')
        self.stdM.place(x=500, y=180)
        click=StringVar()
        click.set("")
        self.drop_medium = OptionMenu(self.root1,click,"" )
        self.drop_medium.place(x=500, y=200)
        self.stdc = Label(self.root1, text="Company Name", font=("Arial", 10, "bold"), bg='#FFFFFF')
        self.stdc.place(x=620, y=180)
        self.drop_bookcn = OptionMenu(self.root1, click, "")
        self.drop_bookcn.place(x=620, y=200)
        self.stdB = Label(self.root1, text="Book Name", font=("Arial", 10, "bold"), bg='#FFFFFF')
        self.stdB.place(x=750, y=180)
        self.drop_bookbn = OptionMenu(self.root1, click, "")
        self.drop_bookbn.place(x=750, y=200)
        self.Cancel = Button(self.root1, text='Cancel', command=self.cancel)
        self.Cancel.place(x=840, y=250)
    def mainfunc(self):
       Title_right = Label(self.root1, text="New Order", font=("Arial", 14, "bold"), fg="Green", bg='#FFFFFF')
       Title_right.place(x=570, y=100)
       New = Button(self.root1,text='New Order', width=10,command=self.new_or)
       New.place(x=450,y=150)
       order_ID = Label(self.root1, text="Order ID", font=("Arial", 8, "bold"), bg='#FFFFFF')
       order_ID.place(x=535, y=150)
       self.search_Text=Entry(self.root1,width=8,bd=3)
       self.search_Text.place(x=590,y=150)
       search_bt= Button(self.root1, text='Open', width=7,command=self.open)
       search_bt.place(x=670, y=150)
       search_bt = Button(self.root1, text='Delete', width=7,command=self.Warning)
       search_bt.place(x=740, y=150)
