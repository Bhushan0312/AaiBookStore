import openpyxl
from tkinter import *
import new_book
print("*************************************************************************")
print("*                    Stating Application .......                        *")
print("*************************************************************************")
class main_code():
    def __init__(self):
        self.xl=openpyxl.load_workbook('Resourse/Book1.xlsx',data_only=True)
        self.sh=self.xl["1"]
        self.root=Tk()
        self.root.title("Aai Book Store")
        self.root.geometry("900x600+50+50")
        canvas = Canvas(bg='#FFFFFF')
        canvas.create_line(400, 200, 400, 500)
        canvas.pack(fill=BOTH, expand=True)
        self.root.configure(bg='#FFFFFF')
        self.root.resizable(False,False)
        self.main_func()
    def error(self):
        self.ex = Label(self.root, text="Please select above Options", fg="red",bg='#FFFFFF')
        self.ex.place(x=50, y=500)

    def medium(self,event):
        try:
            sh = self.xl[event]
            medium_options = ["Other"]
            for i in range(2, (sh.max_row) + 1):
                medium_options.append(sh.cell(i, 1).value)
            medium_options = list(dict.fromkeys(medium_options))
            medium_options.reverse()
            self.click_mdm = StringVar()
            self.click_mdm.set("")
            drop_medium = OptionMenu(self.root, self.click_mdm, *medium_options, command=self.get_std)
            drop_medium.config(width=15)
            drop_medium.place(x=50, y=250)
        except:
            self.error()

    def get_std(self, event):
        self.med_stock = event
        try:
            if event == "Other":
                self.tb_med = Entry(self.root, width=10, bd=3)
                self.tb_med.place(x=200, y=250)
            else:
                pass
            self.med = self.click_mdm.get()
            sh = self.xl[self.click_std.get()]
            slist = ["Other"]
            for i in range(2, (sh.max_row) + 1):
                if event == (sh.cell(i, 1).value):
                    slist.append(sh.cell(i, 3).value)
                else:
                    pass
            slist = list(dict.fromkeys(slist))
            slist.reverse()
            self.click_sub = StringVar()
            self.click_sub.set("")
            self.drop_subject = OptionMenu(self.root, self.click_sub, *slist, command=self.company_name)
            self.drop_subject.config(width=15)
            self.drop_subject.place(x=50, y=310)
        except:
           self.error()

    def company_name(self, event):
        try:
            if event == "Other":
                self.tb_sub = Entry(self.root, width=10, bd=3)
                self.tb_sub.place(x=200, y=310)

            else:
                pass
            self.sub = event
            sh = self.xl[self.click_std.get()]
            cn = ["Other"]
            for i in range(2, (sh.max_row) + 1):
                cn.append(sh.cell(i, 2).value)
            cn = list(dict.fromkeys(cn))
            cn.reverse()
            self.click_cn = StringVar()
            self.click_cn.set("")
            self.drop_cn = OptionMenu(self.root, self.click_cn, *cn, command=self.bookname)
            self.drop_cn.config(width=15)
            self.drop_cn.place(x=50, y=370)
        except:
            self.error()

    def bookname(self, event):
        self.cn=event
        try:
            self.cn = event
            if event == "Other":
                self.tb_cn = Entry(self.root, width=10, bd=3)
                self.tb_cn.place(x=200, y=370)
            else:
                pass
            sh = self.xl[self.click_std.get()]
            blist = ["Other"]
            for i in range(2, (sh.max_row) + 1):
                if event == sh.cell(i, 2).value:
                    blist.append(sh.cell(i, 4).value)
                else:
                    pass
            bn = list(dict.fromkeys(blist))
            bn.reverse()
            self.click_cn = StringVar()
            self.click_cn.set("")
            self.drop_bookcn = OptionMenu(self.root, self.click_cn, *bn, command=self.book_n)
            self.drop_bookcn.config(width=15)
            self.drop_bookcn.place(x=50, y=430)
        except:
            self.error()

    def book_n(self, event):
        self.cmn = event
        if event == "Other":
            self.tb_bn = Entry(self.root, width=10, bd=3)
            self.tb_bn.place(x=200, y=430)
        else:
            pass
    def sold(self):
        row=self.row
        sh = self.xl[self.click_std.get()]
        selling_stock = int(self.tx.get("1.0",END))
        available_stock = int(sh.cell(row, 5).value)
        updated_stock = available_stock - selling_stock
        sh.cell(row, 5).value= updated_stock
        self.xl.save("Resourse/Book1.xlsx")
        self.clear()
        self.quantity2 = Label(self.root, text="Sold. Avilable stock: " + str(sh.cell(row, 5).value), font=("Arial", 10, "bold"),bg='#FFFFFF')
        self.quantity2.place(x=280, y=490)
        self.stock()

    def add(self):
        row = self.row
        sh = self.xl[self.click_std.get()]
        add_quant=int(self.tx1.get("1.0",END))
        available_stock = int(sh.cell(row, 5).value)
        updated_stock=available_stock+add_quant
        sh.cell(row, 5).value = int(updated_stock)
        self.xl.save("Resourse/Book1.xlsx")
        self.clear()
        self.quantity1 = Label(self.root, text=" Updated stock: " + str(sh.cell(row, 5).value), font=("Arial", 10, "bold"),bg='#FFFFFF')
        self.quantity1.place(x=280, y=520)
        self.stock()
    def updatemrp(self):
        self.clear()
        sh=self.xl[self.click_std.get()]
        row = self.row
        sh.cell(row, 6).value = int(self.tx2.get("1.0",END))
        self.xl.save("Resourse/Book1.xlsx")
        self.mrp = Label(self.root, text=" MRP Updated " , font=("Arial", 10, "bold"),bg='#FFFFFF')
        self.mrp.place(x=280, y=550)
        self.stock()
    def Add_stock(self):
        try:
            sh = self.xl[self.click_std.get()]
            flag=0
            for i in range(2, (sh.max_row) + 1):
                if (self.med_stock == sh.cell(i, 1).value and self.cn == sh.cell(i,2).value and self.sub == sh.cell(i, 3).value and self.cmn == sh.cell(i, 4).value):
                    self.clear()
                    quantity = Label(self.root, text="Book has already added",font=("Arial", 10, "bold"),bg='#FFFFFF')
                    quantity.place(x=600, y=560)
                    flag=1
                else:
                    pass
            if flag !=1:
                row= sh.max_row+1
                if self.click_mdm.get() != "Other":
                    sh.cell(row, 1).value = self.click_mdm.get()
                else:
                    if self.tb_med.get() != '':
                        sh.cell(row, 1).value = self.tb_med.get()
                    else:
                        self.error()
                if self.cn != "Other":
                    sh.cell(row, 2).value = self.cn
                else:
                    if self.tb_cn.get() != '':
                        sh.cell(row, 2).value = self.tb_cn.get()
                    else:
                        self.error()
                if self.click_sub.get() != "Other":
                    sh.cell(row, 3).value = self.click_sub.get()
                else:
                    if self.tb_sub.get() != '':
                        sh.cell(row, 3).value = self.tb_sub.get()
                    else:
                        self.error()
                if self.cmn != "Other":
                    sh.cell(row, 4).value = self.cmn
                else:
                    if self.tb_bn.get() != '':
                        sh.cell(row, 4).value = self.tb_bn.get()
                    else:
                        self.error()
                try:
                    sh.cell(row, 5).value = int(self.tbQ.get("1.0", END))
                except:
                    sh.cell(row, 5).value = 0
                try:
                    sh.cell(row, 6).value = int(self.tbP.get("1.0", END))
                except:
                    sh.cell(row, 6).value = 0
                self.clear()
                self.xl.save("Resourse/Book1.xlsx")
                self.update1 = Label(self.root, text="Data Updated Successfully", font=("Arial", 10, "bold"), bg='#FFFFFF')
                self.update1.place(x=150, y=530)
        except:
            self.error()
    def stock(self):
        try:
            self.clear()
            flag=0
            sh=self.xl[self.click_std.get()]
            for i in range(2,(sh.max_row)+1):
                if(self.med_stock == sh.cell(i, 1).value and self.cn == sh.cell(i,2).value and self.sub == sh.cell(i, 3).value and self.cmn == sh.cell(i, 4).value):
                    flag=1
                    self.row = i
                    self.quantity = Label(self.root, text="Available stock: "+str(sh.cell(i,5).value), font=("Arial", 10, "bold"),bg='#FFFFFF')
                    self.quantity.place(x=150, y=460)
                    self.price = Label(self.root, text="Book MRP= " + str(sh.cell(i, 6).value),font=("Arial", 10, "bold"),bg='#FFFFFF')
                    self.price.place(x=270, y=460)
                    self.tx1 = Text(self.root, height=1, width=3,bd=3)
                    self.tx1.insert(END, "1")
                    self.tx1.place(x=150, y=520)
                    self.bt1 = Button(self.root, text="Add Stock", width=10, command=self.add)
                    self.bt1.place(x=200, y=520)
                    self.tx2 = Text(self.root, height=1, width=3,bd=3)
                    self.tx2.place(x=150, y=550)
                    self.bt2 = Button(self.root, text="Update MRP", width=10, command=self.updatemrp)
                    self.bt2.place(x=200, y=550)
                    if (int(sh.cell(i,5).value) >0):
                        self.av=sh.cell(i,5).value
                        self.tx=Text(self.root,height=1,width=3,bd=3)
                        self.tx.insert(END,"1")
                        self.tx.place(x=150,y=490)
                        self.bt = Button(self.root, text="Sold", width=10,command=self.sold)
                        self.bt.place(x=200, y=490)



                else:
                    pass
            if (self.med_stock != "" and self.cn != "" and self.sub != "" and self.cmn != ""):
                if(flag!=1):
                    self.clear()
                    self.bt_add=Button(self.root,text="ADD stock" , width=10,command=self.Add_stock)
                    self.bt_add.place(x=50,y=570)
                    self.std_qn = Label(self.root, text="Quantity:", font=("Arial", 10, "bold"), bg='#FFFFFF')
                    self.std_qn.place(x=50, y=510)
                    self.tbQ = Text(self.root, height=1, width=5,bd=3)
                    self.tbQ.place(x=120, y=510)
                    self.std_prc = Label(self.root, text="Price: ", font=("Arial", 10, "bold"), bg='#FFFFFF')
                    self.std_prc.place(x=50, y=540)
                    self.tbP = Text(self.root, height=1, width=5,bd=3)
                    self.tbP.place(x=120, y=540)


            else:
                self.error()
        except:
           self.error()
    def clear(self):
        try:
            self.quantity.destroy()
            self.price.destroy()
        except:
            pass
        try:
            self.ex.destroy()
        except:
            pass
        try:
            self.quantity1.destroy()
        except:
            pass
        try:
            self.quantity2.destroy()
        except:
            pass
        try:
            self.mrp.destroy()
        except:
            pass
        try:
            self.tx.destroy()
            self.bt.destroy()
        except:
            pass
        try:
            self.update1.destroy()
        except:
            pass
        try:
            self.bt_add.destroy()
            self.std_qn.destroy()
            self.tbQ.destroy()
            self.tbP.destroy()
            self.std_prc.destroy()
        except:
            pass
        try:
            self.tx1.destroy()
            self.tx2.destroy()
            self.bt1.destroy()
            self.bt2.destroy()
        except:
            pass

    def main_func(self):
        logo_img = PhotoImage(file='icons/book.png')
        logo = Label(self.root,image=logo_img,borderwidth=0)
        logo.place(x=305,y=0)

        Title_left = Label(self.root,text="Check / ADD Stock",font=("Arial",14,"bold"),fg="Green",bg='#FFFFFF')
        Title_left.place(x=50,y=100)
        standard_options=[""]
        for i in self.xl.sheetnames:
            standard_options.append(i)

        self.click_std = StringVar()
        self.click_std.set(standard_options[0])
        stdL=Label(self.root,text="Standard",font=("Arial",10,"bold"),bg='#FFFFFF')
        stdL.place(x=50,y=150)
        drop_std=OptionMenu(self.root,self.click_std,*standard_options,command=self.medium)
        drop_std.config(width = 15)
        drop_std.place(x=50, y=180)
        click = StringVar()
        click.set("")
        std_mdm=Label(self.root,text="Medium / Stream of Instruction",font=("Arial",10,"bold"),bg='#FFFFFF')
        std_mdm.place(x=50,y=220)
        drop_medium=OptionMenu(self.root,click,"",command=self.get_std)
        drop_medium.config(width = 15)
        drop_medium.place(x=50, y=250)

        std_sub = Label(self.root, text="Subject", font=("Arial", 10, "bold"),bg='#FFFFFF')
        std_sub.place(x=50, y=285)
        drop_subject = OptionMenu(self.root,click,"")
        drop_subject.place(x=50, y=310)

        std_cn = Label(self.root, text="Company Name", font=("Arial", 10, "bold"),bg='#FFFFFF')
        std_cn.place(x=50, y=345)
        drop_booknm = OptionMenu(self.root, click, "")
        drop_booknm.config(width=15)
        drop_booknm.place(x=50, y=370)
        std_bn = Label(self.root, text="Book Name", font=("Arial", 10, "bold"),bg='#FFFFFF')
        std_bn.place(x=50, y=400)
        click=StringVar()
        click.set("")
        drop_cn = OptionMenu(self.root,click ,"")
        drop_cn.place(x=50, y=430)
        bt= Button(self.root,text="Check Stock",width=10,command=self.stock)
        bt.place(x=50,y=470)
        new_book.new_order(self.root,self.xl)
        self.clear()
        self.root.mainloop()
main_code()
print("*************************************************************************")
print("*                    Closing Application .......                        *")
print("*************************************************************************")