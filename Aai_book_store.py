import openpyxl
from tkinter import *
import new_book
print("*************************************************************************")
print("*                    Stating Application .......                        *")
print("*************************************************************************")
class main_code():
    def __init__(self):
        self.xl=openpyxl.load_workbook('Resourse/Book1.xlsx',data_only=True)
        self.sh=self.xl["1"]
        self.root=Tk()
        self.root.title("Aai Book Store")
        self.root.geometry("800x600+50+50")
        canvas = Canvas(bg='#FFFFFF')
        canvas.create_line(400, 200, 400, 500)
        canvas.pack(fill=BOTH, expand=True)
        self.root.configure(bg='#FFFFFF')
        self.root.resizable(False,False)
        self.main_func()
    def error(self):
        self.ex = Label(self.root, text="Please select above Options", fg="red",bg='#FFFFFF')
        self.ex.place(x=50, y=500)

    def medium(self,event):
        sh=self.xl[event]
        medium_options = [""]
        for i in range(2, (sh.max_row) + 1):
            medium_options.append(sh.cell(i, 1).value)
        medium_options = list(dict.fromkeys(medium_options))
        self.click_mdm = StringVar()
        self.click_mdm.set(medium_options[0])
        drop_medium = OptionMenu(self.root, self.click_mdm, *medium_options, command=self.get_std)
        drop_medium.config(width=15)
        drop_medium.place(x=50, y=250)
    def get_std(self,event):
        try:
            sh=self.xl[self.click_std.get()]
            slist = []
            for i in range(2,(sh.max_row)+1):
                if event == (sh.cell(i,1).value):
                    slist.append(sh.cell(i, 3).value)
                elif(event=="Both"):
                    slist.append(sh.cell(i, 3).value)

            slist=list(dict.fromkeys(slist))
            click_sub = StringVar()
            click_sub.set(slist[0])
            self.drop_subject = OptionMenu(self.root, click_sub, *slist,command=self.company_name)
            self.drop_subject.config(width=15)
            self.drop_subject.place(x=50, y=310)
            self.sub=click_sub.get()
        except:
            self.error()

    def company_name(self,event):
        self.sub=event
        sh=self.xl[self.click_std.get()]
        cn = [""]
        for i in range(2, (sh.max_row) + 1):
            cn.append(sh.cell(i, 2).value)
        cn = list(dict.fromkeys(cn))
        click_cn = StringVar()
        drop_booknm = OptionMenu(self.root, click_cn, *cn, command=self.bookname)
        drop_booknm.config(width=15)
        drop_booknm.place(x=50, y=370)
        self.cget=click_cn.get()
    def get_bn(self,event):
        self.conm=event

    def bookname(self,event):
            self.cn=event
            sh=self.xl[self.click_std.get()]
            blist=[]
            for i in range(2,(sh.max_row)+1):
                if event ==sh.cell(i,2).value:
                    blist.append(sh.cell(i,4).value)
                else:
                   pass
            bn=list(dict.fromkeys(blist))
            self.click_cn=StringVar()
            try:
               self.click_cn.set(blist[0])
            except:
                self.clear()
            try:
                self.drop_bookcn = OptionMenu(self.root, self.click_cn,*bn,command=self.get_bn)
                self.drop_bookcn.config(width=15)
                self.drop_bookcn.place(x=50, y=430)
                self.get_bn(self.click_cn.get())
            except:
                pass
    def sold(self):
        row=self.row
        sh = self.xl[self.click_std.get()]
        selling_stock = int(self.tx.get("1.0",END))
        available_stock = int(sh.cell(row, 5).value)
        updated_stock = available_stock - selling_stock
        sh.cell(row, 5).value= updated_stock
        self.xl.save("Resourse/Book1.xlsx")
        self.clear()
        self.quantity2 = Label(self.root, text="Sold. Avilable stock: " + str(sh.cell(row, 5).value), font=("Arial", 10, "bold"),bg='#FFFFFF')
        self.quantity2.place(x=280, y=490)
        self.stock()

    def add(self):
        row = self.row
        sh = self.xl[self.click_std.get()]
        add_quant=int(self.tx1.get("1.0",END))
        available_stock = int(sh.cell(row, 5).value)
        updated_stock=available_stock+add_quant
        sh.cell(row, 5).value = int(updated_stock)
        self.xl.save("Resourse/Book1.xlsx")
        self.clear()
        self.quantity1 = Label(self.root, text=" Updated stock: " + str(sh.cell(row, 5).value), font=("Arial", 10, "bold"),bg='#FFFFFF')
        self.quantity1.place(x=280, y=520)
        self.stock()
    def updatemrp(self):
        self.clear()
        sh=self.xl[self.click_std.get()]
        row = self.row
        sh.cell(row, 6).value = int(self.tx2.get("1.0",END))
        self.xl.save("Resourse/Book1.xlsx")
        self.mrp = Label(self.root, text=" MRP Updated " , font=("Arial", 10, "bold"),bg='#FFFFFF')
        self.mrp.place(x=280, y=550)
        self.stock()
    def stock(self):
        try:
            self.clear()
            sh=self.xl[self.click_std.get()]
            for i in range(2,(sh.max_row)+1):
                if(self.click_mdm.get()==sh.cell(i,1).value and self.cn==sh.cell(i,2).value and self.sub==sh.cell(i,3).value and self.conm == sh.cell(i,4).value):
                    self.row = i
                    self.quantity = Label(self.root, text="Avilable stock: "+str(sh.cell(i,5).value), font=("Arial", 10, "bold"),bg='#FFFFFF')
                    self.quantity.place(x=150, y=460)
                    self.price = Label(self.root, text="Book MRP= " + str(sh.cell(i, 6).value),font=("Arial", 10, "bold"),bg='#FFFFFF')
                    self.price.place(x=270, y=460)
                    self.tx1 = Text(self.root, height=1, width=3,bd=3)
                    self.tx1.insert(END, "1")
                    self.tx1.place(x=150, y=520)
                    self.bt1 = Button(self.root, text="Add Stock", width=10, command=self.add)
                    self.bt1.place(x=200, y=520)
                    self.tx2 = Text(self.root, height=1, width=3,bd=3)
                    self.tx2.place(x=150, y=550)
                    self.bt2 = Button(self.root, text="Update MRP", width=10, command=self.updatemrp)
                    self.bt2.place(x=200, y=550)
                    if (int(sh.cell(i,5).value) >0):
                        self.av=sh.cell(i,5).value
                        self.tx=Text(self.root,height=1,width=3,bd=3)
                        self.tx.insert(END,"1")
                        self.tx.place(x=150,y=490)
                        self.bt = Button(self.root, text="Sold", width=10,command=self.sold)
                        self.bt.place(x=200, y=490)



                else:
                    pass
        except:
           self.error()
    def clear(self):
        try:
            self.quantity.destroy()
            self.price.destroy()
        except:
            pass
        try:
            self.std_bn.destroy()
        except:
            pass
        try:
            self.ex.destroy()
        except:
            pass
        try:
            self.quantity1.destroy()
        except:
            print("pass add")
        try:
            self.quantity2.destroy()
        except:
            pass
        try:
            self.mrp.destroy()
        except:
            pass
        try:
            self.tx.destroy()
            self.bt.destroy()
        except:
            pass
    def call_class(self):
        new_book.add_book(self.root,self.xl)

    def main_func(self):
        logo_img = PhotoImage(file='icons/book.png')
        logo = Label(self.root,image=logo_img,borderwidth=0)
        logo.place(x=350,y=0)
        Title_left = Label(self.root,text="Check Stock",font=("Arial",14,"bold"),fg="Green",bg='#FFFFFF')
        Title_left.place(x=50,y=100)
        standard_options=[""]
        for i in self.xl.sheetnames:
            standard_options.append(i)

        self.click_std = StringVar()
        self.click_std.set(standard_options[0])
        stdL=Label(self.root,text="Standard",font=("Arial",10,"bold"),bg='#FFFFFF')
        stdL.place(x=50,y=150)
        drop_std=OptionMenu(self.root,self.click_std,*standard_options,command=self.medium)
        drop_std.config(width = 15)
        drop_std.place(x=50, y=180)
        click = StringVar()
        click.set("")
        std_mdm=Label(self.root,text="Medium / Stream of Instruction",font=("Arial",10,"bold"),bg='#FFFFFF')
        std_mdm.place(x=50,y=220)
        drop_medium=OptionMenu(self.root,click,"",command=self.get_std)
        drop_medium.config(width = 15)
        drop_medium.place(x=50, y=250)

        std_sub = Label(self.root, text="Subject", font=("Arial", 10, "bold"),bg='#FFFFFF')
        std_sub.place(x=50, y=285)
        drop_subject = OptionMenu(self.root,click,"")
        drop_subject.place(x=50, y=310)

        std_cn = Label(self.root, text="Company Name", font=("Arial", 10, "bold"),bg='#FFFFFF')
        std_cn.place(x=50, y=345)
        drop_booknm = OptionMenu(self.root, click, "")
        drop_booknm.config(width=15)
        drop_booknm.place(x=50, y=370)
        std_bn = Label(self.root, text="Book Name", font=("Arial", 10, "bold"),bg='#FFFFFF')
        std_bn.place(x=50, y=400)
        click=StringVar()
        click.set("")
        drop_cn = OptionMenu(self.root,click ,"")
        drop_cn.place(x=50, y=430)
        bt= Button(self.root,text="Check Stock",width=10,command=self.stock)
        bt.place(x=50,y=470)
        new_book.add_book(self.root, self.xl)
        self.clear()
        self.root.mainloop()
main_code()
print("*************************************************************************")
print("*                    Closing Application .......                        *")
print("*************************************************************************")